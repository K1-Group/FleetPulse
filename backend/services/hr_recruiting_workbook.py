"""Read-only HR recruiting KPI workbook projection.

The workbook parser only projects source-backed aggregates and masked exception
queue evidence. It never returns applicant names, phone numbers, or email
addresses to FleetPulse.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta, timezone
import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from utils.dashboard_date_range import DashboardDateRange, dashboard_date_range, pct_change

SOURCE_PROFILE = "kpi_workbook"
SOURCE_SYSTEM = "HR Lead KPI Recheck workbook"
SOURCE_AUTHORITY = "Grasshopper/SharePoint/Tenstreet HR KPI recheck workbook"
CONVERSION_SOURCE_SYSTEM = "HR lead-to-Xcelerator conversion workbook"
CONVERSION_SOURCE_AUTHORITY = "Xcelerator driver list exact-name conversion workbook"

EXPECTED_TABS = (
    "Lead Level KPI",
    "Call Attempts Detail",
    "Failed No Outbound",
    "Recovered 24-48h",
    "Failed Over 72h",
    "No Real Discussion",
    "Source Log QA",
)
CONVERSION_EXPECTED_TABS = (
    "Summary",
    "Lead Conversion Detail",
    "Converted Leads",
    "Not Converted Leads",
    "Conversion by Source",
    "Rules and Issues",
)

FIRST_OUTREACH_BUCKET_ORDER = (
    "Within 24h",
    "24-48h recovered",
    "48-72h late",
    "Over 72h failed",
    "No outbound found",
)
REAL_DISCUSSION_BUCKET_ORDER = (
    "Within 24h",
    "24-48h recovered",
    "48-72h late",
    "Over 72h failed",
    "No 1min+ discussion found",
)
EXCEPTION_TABS = (
    "Failed No Outbound",
    "Failed Over 72h",
    "No Real Discussion",
)
EXCEPTION_TAB_SPECS = {
    "Failed No Outbound": {
        "category": "Failed No Outbound",
        "reason": "No outbound found",
        "severity": "critical",
    },
    "Failed Over 72h": {
        "category": "Failed Over 72h",
        "reason": "Over 72h failed",
        "severity": "critical",
    },
    "No Real Discussion": {
        "category": "No Real Discussion",
        "reason": "No 1min+ discussion found",
        "severity": "warning",
    },
}
LEAD_CREATED_ALIASES = (
    "Lead Created At",
    "Created At",
    "Submitted At",
    "Application Date",
    "Date",
)
STATUS_ALIASES = (
    "App Status",
    "Status",
    "Application Status",
    "Lead Status",
)
FIRST_OUTREACH_BUCKET_ALIASES = (
    "First Outreach KPI Bucket",
    "First Outreach Bucket",
)
REAL_DISCUSSION_BUCKET_ALIASES = (
    "Real Discussion KPI Bucket",
    "Real Discussion Bucket",
)
IDENTITY_ALIASES = (
    "Lead ID",
    "Lead Form ID",
    "Application ID",
    "Applicant ID",
    "Lead Name",
    "Name",
    "Applicant",
    "Candidate",
    "Driver",
    "Phone",
    "Phone Number",
    "Email",
)
OUTREACH_HOUR_ALIASES = (
    "Hours to First Outreach",
    "Hours To First Outreach",
    "First Outreach Hours",
)
DISCUSSION_HOUR_ALIASES = (
    "Hours to First Real Discussion",
    "Hours To First Real Discussion",
    "First Real Discussion Hours",
)
LEAD_INTAKE_DATE_FIELDS = ("Lead Created At", "Submitted At", "Created At")
LEAD_OUTCOME_DATE_FIELDS = ("Interview Scheduled At", "Orientation Scheduled At", "Hire Date", "Hired At", "Completed At")
CALL_ATTEMPT_DATE_FIELDS = ("Call Date/Time", "Created At")
CONVERSION_INTAKE_DATE_FIELDS = ("Lead Created At", "Application Date")
CONVERSION_CONVERTED_ALIASES = ("Converted By Rule", "Converted")
CONVERSION_STILL_DRIVING_ALIASES = ("Still Driving Evidence", "Still Driving")
CONVERSION_PHONE_MATCH_ALIASES = ("Phone Also Matches Driver", "Phone Match")
CONVERSION_SOURCE_ALIASES = ("Source Systems", "Source")
CONVERSION_SLA_ALIASES = ("SLA Result", "First Outreach SLA")


def build_hr_recruiting_workbook_dataset(
    workbook_path: str,
    *,
    now: datetime,
    source: str,
    table_id: str,
    sla_hours: tuple[int, ...],
    hard_targets: dict[str, dict[str, Any]],
    start_date: date | str | None = None,
    end_date: date | str | None = None,
    conversion_workbook_path: str = "",
) -> dict[str, Any]:
    """Build the HR recruiting dataset from the KPI recheck workbook."""

    path = Path(workbook_path).expanduser()
    selected_range = dashboard_date_range(start_date, end_date)
    conversion_funnel = _conversion_funnel_from_path(conversion_workbook_path, selected_range)
    if not workbook_path:
        return _empty_dataset(
            now=now,
            source=source,
            table_id=table_id,
            sla_hours=sla_hours,
            hard_targets=hard_targets,
            source_status="snapshot_not_configured",
            source_message="Configure HR_RECRUITING_WORKBOOK_PATH with the approved HR KPI recheck workbook.",
            selected_range=selected_range,
            conversion_funnel=conversion_funnel,
        )
    if not path.exists():
        return _empty_dataset(
            now=now,
            source=source,
            table_id=table_id,
            sla_hours=sla_hours,
            hard_targets=hard_targets,
            source_status="source_error",
            source_message="Configured HR_RECRUITING_WORKBOOK_PATH does not exist.",
            workbook_name=path.name,
            selected_range=selected_range,
            conversion_funnel=conversion_funnel,
        )

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - keep dashboard status available.
        return _empty_dataset(
            now=now,
            source=source,
            table_id=table_id,
            sla_hours=sla_hours,
            hard_targets=hard_targets,
            source_status="source_error",
            source_message=f"HR KPI workbook could not be opened: {type(exc).__name__}.",
            workbook_name=path.name,
            selected_range=selected_range,
            conversion_funnel=conversion_funnel,
        )

    present_tabs = [sheet for sheet in EXPECTED_TABS if sheet in workbook.sheetnames]
    missing_tabs = [sheet for sheet in EXPECTED_TABS if sheet not in workbook.sheetnames]
    if "Lead Level KPI" not in workbook.sheetnames:
        return _empty_dataset(
            now=now,
            source=source,
            table_id=table_id,
            sla_hours=sla_hours,
            hard_targets=hard_targets,
            source_status="source_error",
            source_message="HR KPI workbook is missing required tab: Lead Level KPI.",
            workbook_name=path.name,
            present_tabs=present_tabs,
            missing_tabs=missing_tabs,
            selected_range=selected_range,
            conversion_funnel=conversion_funnel,
        )

    leads = _records_for_sheet(
        workbook,
        "Lead Level KPI",
        required_headers=("Lead Created At", "First Outreach KPI Bucket", "Real Discussion KPI Bucket"),
    )
    call_attempts = _records_for_sheet(
        workbook,
        "Call Attempts Detail",
        required_headers=("Call Date/Time", "HR Member", "Duration Seconds"),
    )
    first_outreach_by_member = _member_kpi_rows(
        _records_for_sheet(workbook, "KPI By First Outreach", required_headers=("HR Member", "Within_24h"))
    )
    real_discussion_by_member = _member_kpi_rows(
        _records_for_sheet(workbook, "KPI By Real Discussion", required_headers=("HR Member", "Within_24h"))
    )
    source_log_qa = _source_log_qa_rows(
        _records_for_sheet(workbook, "Source Log QA", required_headers=("File", "Rows", "Used for Mapping"))
    )
    exception_records_by_tab = {
        sheet: _filter_records_by_date_range(
            _records_for_sheet(
                workbook,
                sheet,
                required_headers=("Lead Created At", "First Outreach KPI Bucket", "Real Discussion KPI Bucket"),
            ),
            selected_range,
            LEAD_INTAKE_DATE_FIELDS,
        )
        for sheet in EXCEPTION_TABS
    }
    unfiltered_leads = leads
    unfiltered_call_attempts = call_attempts
    leads = _filter_records_by_date_range(
        leads,
        selected_range,
        LEAD_INTAKE_DATE_FIELDS,
    )
    call_attempts = _filter_records_by_date_range(
        call_attempts,
        selected_range,
        CALL_ATTEMPT_DATE_FIELDS,
    )

    as_of = _ensure_aware(now)
    if not unfiltered_leads:
        source_status = "empty"
        source_message = "HR KPI workbook is present but Lead Level KPI contains no rows."
    elif selected_range and not leads:
        source_status = "empty"
        source_message = "HR KPI workbook is present but no Lead Level KPI rows matched the selected date range."
    else:
        source_status = "partial" if missing_tabs else "ok"
        source_message = (
            f"HR KPI workbook is missing optional evidence tab(s): {', '.join(missing_tabs)}."
            if missing_tabs
            else None
        )

    first_outreach_counts = Counter(_text(row.get("First Outreach KPI Bucket"), "Unknown") for row in leads)
    real_discussion_counts = Counter(_text(row.get("Real Discussion KPI Bucket"), "Unknown") for row in leads)
    status_counts = Counter(_text(row.get("App Status"), "No Status") for row in leads)
    lead_count = len(leads)
    within_24h = first_outreach_counts.get("Within 24h", 0)
    recovered_24_48 = first_outreach_counts.get("24-48h recovered", 0)
    late_48_72 = first_outreach_counts.get("48-72h late", 0)
    failed_over_72 = first_outreach_counts.get("Over 72h failed", 0)
    no_outbound = first_outreach_counts.get("No outbound found", 0)
    no_real_discussion = real_discussion_counts.get("No 1min+ discussion found", 0)
    first_touch_pct = round(within_24h / lead_count, 4) if lead_count else None
    first_outreach_hours = [
        value
        for value in (_number(row.get("Hours to First Outreach")) for row in leads)
        if value is not None
    ]
    real_discussion_hours = [
        value
        for value in (_number(row.get("Hours to First Real Discussion")) for row in leads)
        if value is not None
    ]
    source_rows_by_tab = _tab_counts(workbook, present_tabs)
    kpi_summary = {
        "unique_lead_forms": lead_count,
        "outbound_within_24h": within_24h,
        "outbound_24_48h": recovered_24_48,
        "outbound_48_72h": late_48_72,
        "outbound_over_72h": failed_over_72,
        "no_outbound_found": no_outbound,
        "any_outbound_after_form": lead_count - no_outbound,
        "no_real_discussion_found": no_real_discussion,
        "real_discussion_within_24h": real_discussion_counts.get("Within 24h", 0),
        "total_outbound_attempts": len(call_attempts),
        "total_1min_discussions": sum(1 for row in call_attempts if _boolish(row.get("Real Discussion 1min+"))),
        "avg_hours_to_first_outreach": _mean(first_outreach_hours),
        "avg_hours_to_first_real_discussion": _mean(real_discussion_hours),
        "first_touch_24h_pct": first_touch_pct,
    }
    exception_queue = _exception_queue_rows(
        leads,
        exception_records_by_tab,
        as_of=as_of,
    )

    summary = {
        "active_leads": lead_count,
        "new_leads_today": (
            _record_count_for_range(leads, selected_range, LEAD_INTAKE_DATE_FIELDS)
            if selected_range
            else sum(1 for row in leads if _same_day(row.get("Lead Created At"), as_of))
        ),
        "avg_process_age_hours": kpi_summary["avg_hours_to_first_outreach"],
        "stale_leads": late_48_72 + failed_over_72 + no_outbound,
        "completed_today": 0,
        "new_hires_7d": 0,
        "active_qualified_pipeline": 0,
        "first_touch_24h_pct": first_touch_pct,
        "first_touch_eligible_count": lead_count,
        "first_touch_within_24h_count": within_24h,
        "stale_untouched_48h": no_outbound,
        "orientation_scheduled_count": 0,
        "orientation_show_count": 0,
        "orientation_show_rate": None,
    }
    hard_target_results = _build_hard_target_results(
        hard_targets,
        {
            "new_hires_7d": None,
            "active_qualified_pipeline": None,
            "first_touch_24h_pct": first_touch_pct,
            "stale_untouched_48h": no_outbound,
            "orientation_show_rate": None,
        },
        {
            "new_hires_7d": False,
            "active_qualified_pipeline": False,
            "first_touch_24h_pct": lead_count > 0,
            "stale_untouched_48h": lead_count > 0,
            "orientation_show_rate": False,
        },
    )
    hard_target_misses = [key for key, target in hard_target_results.items() if target["status"] == "warning"]
    hard_target_pending = [key for key, target in hard_target_results.items() if target["status"] == "awaiting_feed"]
    hard_target_status = (
        "awaiting_feed"
        if hard_target_pending and not hard_target_misses
        else "healthy" if not hard_target_misses and not hard_target_pending else "warning"
    )

    return {
        "generated_at": as_of.isoformat(),
        "projection_mode": "read_only",
        "source_profile": SOURCE_PROFILE,
        "source_system": SOURCE_SYSTEM,
        "source_authority": SOURCE_AUTHORITY,
        "source": source,
        "table_id": table_id,
        "source_artifact": path.name,
        "source_status": source_status,
        "source_message": source_message,
        "pii_suppressed": True,
        "sla_hours": list(sla_hours),
        "hard_targets": hard_target_results,
        "hard_target_status": hard_target_status,
        "hard_target_misses": hard_target_misses,
        "hard_target_pending": hard_target_pending,
        "date_range": selected_range.as_dict() if selected_range else None,
        "period_metrics": _period_metrics(leads, call_attempts, selected_range),
        "trend_comparison": _trend_comparison(unfiltered_leads, unfiltered_call_attempts, selected_range),
        "summary": summary,
        "by_worklist": _first_outreach_bucket_rows(leads, first_outreach_counts),
        "daily": _daily_rows(leads),
        "status_counts": [
            {"status": status, "count": count}
            for status, count in sorted(status_counts.items(), key=lambda item: (-item[1], item[0]))
        ],
        "trend": _trend_rows(leads),
        "row_counts": {
            "source_rows": len(unfiltered_leads),
            "deduped_leads": lead_count,
            "unfiltered_deduped_leads": len(unfiltered_leads),
            "active_leads": lead_count,
            "completed_leads": 0,
            "invalid_rows": 0,
            "source_email_dedupe_rows": 0,
            "call_attempt_rows": len(call_attempts),
            "unfiltered_call_attempt_rows": len(unfiltered_call_attempts),
            "source_log_qa_rows": len(source_log_qa),
            "exception_queue_rows": len(exception_queue),
            "tabs_present": len(present_tabs),
            "tabs_missing": len(missing_tabs),
            "conversion_source_rows": _conversion_count(conversion_funnel, "unfiltered_eligible_leads"),
            "conversion_eligible_leads": _conversion_count(conversion_funnel, "eligible_leads"),
            "conversion_converted_leads": _conversion_count(conversion_funnel, "converted_leads"),
        },
        "validation_errors": {"missing_tabs": len(missing_tabs)} if missing_tabs else {},
        "workbook_evidence": {
            "workbook_name": path.name,
            "tabs": source_rows_by_tab,
            "missing_tabs": missing_tabs,
            "kpi_summary": kpi_summary,
            "first_outreach_buckets": _bucket_counts(first_outreach_counts, FIRST_OUTREACH_BUCKET_ORDER),
            "real_discussion_buckets": _bucket_counts(real_discussion_counts, REAL_DISCUSSION_BUCKET_ORDER),
            "first_outreach_by_member": [] if selected_range else first_outreach_by_member,
            "real_discussion_by_member": [] if selected_range else real_discussion_by_member,
            "source_log_qa": source_log_qa,
            "exception_queue": exception_queue,
            "exception_summary": _exception_summary(exception_queue),
            "conversion_funnel": conversion_funnel,
        },
    }


def _empty_dataset(
    *,
    now: datetime,
    source: str,
    table_id: str,
    sla_hours: tuple[int, ...],
    hard_targets: dict[str, dict[str, Any]],
    source_status: str,
    source_message: str,
    workbook_name: str | None = None,
    present_tabs: list[str] | None = None,
    missing_tabs: list[str] | None = None,
    selected_range: DashboardDateRange | None = None,
    conversion_funnel: dict[str, Any] | None = None,
) -> dict[str, Any]:
    as_of = _ensure_aware(now)
    hard_target_results = _build_hard_target_results(
        hard_targets,
        {key: None for key in hard_targets},
        {key: False for key in hard_targets},
    )
    return {
        "generated_at": as_of.isoformat(),
        "projection_mode": "read_only",
        "source_profile": SOURCE_PROFILE,
        "source_system": SOURCE_SYSTEM,
        "source_authority": SOURCE_AUTHORITY,
        "source": source,
        "table_id": table_id,
        "source_artifact": workbook_name,
        "source_status": source_status,
        "source_message": source_message,
        "pii_suppressed": True,
        "sla_hours": list(sla_hours),
        "hard_targets": hard_target_results,
        "hard_target_status": "awaiting_feed",
        "hard_target_misses": [],
        "hard_target_pending": list(hard_targets),
        "date_range": selected_range.as_dict() if selected_range else None,
        "period_metrics": {
            "new_leads": 0,
            "new_applicants": 0,
            "interviews_scheduled": 0,
            "new_hires": 0,
            "follow_ups": 0,
        },
        "trend_comparison": None,
        "summary": {
            "active_leads": 0,
            "new_leads_today": 0,
            "avg_process_age_hours": 0,
            "stale_leads": 0,
            "completed_today": 0,
            "new_hires_7d": 0,
            "active_qualified_pipeline": 0,
            "first_touch_24h_pct": None,
            "first_touch_eligible_count": 0,
            "first_touch_within_24h_count": 0,
            "stale_untouched_48h": 0,
            "orientation_scheduled_count": 0,
            "orientation_show_count": 0,
            "orientation_show_rate": None,
        },
        "by_worklist": [],
        "daily": [],
        "status_counts": [],
        "trend": [],
        "row_counts": {
            "source_rows": 0,
            "deduped_leads": 0,
            "unfiltered_deduped_leads": 0,
            "active_leads": 0,
            "completed_leads": 0,
            "invalid_rows": 0,
            "source_email_dedupe_rows": 0,
            "call_attempt_rows": 0,
            "unfiltered_call_attempt_rows": 0,
            "source_log_qa_rows": 0,
            "exception_queue_rows": 0,
            "tabs_present": len(present_tabs or []),
            "tabs_missing": len(missing_tabs or EXPECTED_TABS),
            "conversion_source_rows": _conversion_count(conversion_funnel, "unfiltered_eligible_leads"),
            "conversion_eligible_leads": _conversion_count(conversion_funnel, "eligible_leads"),
            "conversion_converted_leads": _conversion_count(conversion_funnel, "converted_leads"),
        },
        "validation_errors": {"missing_tabs": len(missing_tabs or EXPECTED_TABS)},
        "workbook_evidence": {
            "workbook_name": workbook_name,
            "tabs": [
                {"sheet": sheet, "row_count": 0, "status": "present"}
                for sheet in (present_tabs or [])
            ],
            "missing_tabs": list(missing_tabs or EXPECTED_TABS),
            "kpi_summary": {},
            "first_outreach_buckets": [],
            "real_discussion_buckets": [],
            "first_outreach_by_member": [],
            "real_discussion_by_member": [],
            "source_log_qa": [],
            "exception_queue": [],
            "exception_summary": [],
            "conversion_funnel": conversion_funnel,
        },
    }


def _records_for_sheet(
    workbook: Workbook,
    sheet_name: str,
    *,
    required_headers: tuple[str, ...],
) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []

    worksheet = workbook[sheet_name]
    header_row_index = _find_header_row(worksheet, required_headers)
    if header_row_index is None:
        return []

    headers = [
        _text(cell)
        for cell in next(worksheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))
    ]
    records: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        record = {
            header: value
            for header, value in zip(headers, row)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def _find_header_row(worksheet: Any, required_headers: tuple[str, ...]) -> int | None:
    required = {_normalize_header(header) for header in required_headers}
    best: tuple[int, int] | None = None
    max_row = min(20, int(worksheet.max_row or 0))
    for index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        normalized = {_normalize_header(_text(value)) for value in row if _text(value)}
        score = len(required.intersection(normalized))
        if score and (best is None or score > best[0]):
            best = (score, index)
    if best and best[0] >= min(2, len(required)):
        return best[1]
    return None


def _tab_counts(workbook: Workbook, present_tabs: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet in present_tabs:
        required = {
            "Source Log QA": ("File", "Rows", "Used for Mapping"),
            "Call Attempts Detail": ("Call Date/Time", "HR Member", "Duration Seconds"),
        }.get(sheet, ("Lead Created At", "First Outreach KPI Bucket", "Real Discussion KPI Bucket"))
        records = _records_for_sheet(workbook, sheet, required_headers=required)
        rows.append({"sheet": sheet, "row_count": len(records), "status": "present"})
    return rows


def _first_outreach_bucket_rows(
    leads: list[dict[str, Any]],
    first_outreach_counts: Counter[str],
) -> list[dict[str, Any]]:
    hours_by_bucket: dict[str, list[float]] = defaultdict(list)
    for row in leads:
        bucket = _text(row.get("First Outreach KPI Bucket"), "Unknown")
        hours = _number(row.get("Hours to First Outreach"))
        if hours is not None:
            hours_by_bucket[bucket].append(hours)

    ordered = _bucket_counts(first_outreach_counts, FIRST_OUTREACH_BUCKET_ORDER)
    rows: list[dict[str, Any]] = []
    for item in ordered:
        bucket = item["bucket"]
        count = item["count"]
        stale_24h = count if bucket in {"24-48h recovered", "48-72h late", "Over 72h failed", "No outbound found"} else 0
        stale_48h = count if bucket in {"48-72h late", "Over 72h failed", "No outbound found"} else 0
        stale_72h = count if bucket in {"Over 72h failed", "No outbound found"} else 0
        rows.append(
            {
                "worklist": bucket,
                "active_leads": count,
                "new_leads_today": 0,
                "avg_age_hours": _mean(hours_by_bucket.get(bucket, [])),
                "max_age_hours": _max_or_zero(hours_by_bucket.get(bucket, [])),
                "stale_24h": stale_24h,
                "stale_48h": stale_48h,
                "stale_72h": stale_72h,
            }
        )
    return rows


def _daily_rows(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {"new_leads": 0, "completed_leads": 0, "active_leads": 0, "process_times": []}
    )
    for row in leads:
        lead_date = _parse_datetime(row.get("Lead Created At"))
        if lead_date is None:
            continue
        bucket = _text(row.get("First Outreach KPI Bucket"), "Unknown")
        key = (lead_date.date().isoformat(), bucket)
        buckets[key]["new_leads"] += 1
        buckets[key]["active_leads"] += 1
        if _text(row.get("Real Discussion KPI Bucket")) not in {"", "No 1min+ discussion found"}:
            buckets[key]["completed_leads"] += 1
        hours = _number(row.get("Hours to First Outreach"))
        if hours is not None:
            buckets[key]["process_times"].append(hours)

    rows: list[dict[str, Any]] = []
    for (day, bucket), values in buckets.items():
        rows.append(
            {
                "date": day,
                "worklist": bucket,
                "new_leads": values["new_leads"],
                "completed_leads": values["completed_leads"],
                "active_leads": values["active_leads"],
                "avg_process_time_hours": _mean(values["process_times"]),
            }
        )
    return sorted(rows, key=lambda row: (str(row["date"]), str(row["worklist"])))


def _trend_rows(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[str, dict[str, Any]] = defaultdict(lambda: {"new_leads": 0, "stale_leads": 0, "ages": []})
    for row in leads:
        lead_date = _parse_datetime(row.get("Lead Created At"))
        if lead_date is None:
            continue
        day = lead_date.date().isoformat()
        buckets[day]["new_leads"] += 1
        bucket = _text(row.get("First Outreach KPI Bucket"))
        if bucket in {"48-72h late", "Over 72h failed", "No outbound found"}:
            buckets[day]["stale_leads"] += 1
        hours = _number(row.get("Hours to First Outreach"))
        if hours is not None:
            buckets[day]["ages"].append(hours)

    rows: list[dict[str, Any]] = []
    for day, values in buckets.items():
        rows.append(
            {
                "date": day,
                "active_leads": values["new_leads"],
                "new_leads": values["new_leads"],
                "stale_leads": values["stale_leads"],
                "avg_age_hours": _mean(values["ages"]),
            }
        )
    return sorted(rows, key=lambda row: str(row["date"]))


def _member_kpi_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        member = _text(record.get("HR Member"))
        if not member:
            continue
        rows.append(
            {
                "hr_member": member,
                "lead_count": _int(record.get("Leads_First_Outreach") or record.get("Leads_First_Real_Discussion")),
                "within_24h": _int(record.get("Within_24h")),
                "recovered_24_48h": _int(record.get("Recovered_24_48")),
                "late_48_72h": _int(record.get("Late_48_72")),
                "failed_over_72h": _int(record.get("Failed_Over_72")),
                "avg_hours": _number(record.get("Avg_Hours_To_First_Outreach") or record.get("Avg_Hours_To_First_Real")),
                "median_hours": _number(record.get("Median_Hours_To_First_Outreach") or record.get("Median_Hours_To_First_Real")),
                "within_24h_rate": _number(record.get("Within_24h_Rate") or record.get("Real_Within_24_Rate")),
                "total_outbound_attempts": _int(record.get("Total_Outbound_Attempts")),
            }
        )
    return rows


def _source_log_qa_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        file_name = _text(record.get("File"))
        if not file_name:
            continue
        rows.append(
            {
                "file": file_name,
                "row_count": _int(record.get("Rows")),
                "column_count": _int(record.get("Columns")),
                "used_for_mapping": _boolish(record.get("Used for Mapping")),
                "notes": _text(record.get("Reason / Notes")),
                "first_columns": _text(record.get("First Columns")),
            }
        )
    return rows


def _conversion_funnel_from_path(
    workbook_path: str,
    selected_range: DashboardDateRange | None,
) -> dict[str, Any] | None:
    if not workbook_path:
        return None

    path = Path(workbook_path).expanduser()
    if not path.exists():
        return _empty_conversion_funnel(
            workbook_name=path.name,
            source_status="source_error",
            source_message="Configured HR_RECRUITING_CONVERSION_WORKBOOK_PATH does not exist.",
            selected_range=selected_range,
        )

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - keep dashboard status available.
        return _empty_conversion_funnel(
            workbook_name=path.name,
            source_status="source_error",
            source_message=f"HR conversion workbook could not be opened: {type(exc).__name__}.",
            selected_range=selected_range,
        )

    present_tabs = [sheet for sheet in CONVERSION_EXPECTED_TABS if sheet in workbook.sheetnames]
    missing_tabs = [sheet for sheet in CONVERSION_EXPECTED_TABS if sheet not in workbook.sheetnames]
    if "Lead Conversion Detail" not in workbook.sheetnames:
        return _empty_conversion_funnel(
            workbook_name=path.name,
            source_status="source_error",
            source_message="HR conversion workbook is missing required tab: Lead Conversion Detail.",
            present_tabs=present_tabs,
            missing_tabs=missing_tabs,
            selected_range=selected_range,
        )

    detail_rows = _records_for_sheet(
        workbook,
        "Lead Conversion Detail",
        required_headers=("Lead Created At", "Source Systems", "Converted By Rule"),
    )
    filtered_rows = _filter_records_by_date_range(
        detail_rows,
        selected_range,
        CONVERSION_INTAKE_DATE_FIELDS,
    )
    source_status = "ok"
    source_message: str | None = None
    if not detail_rows:
        source_status = "empty"
        source_message = "HR conversion workbook is present but Lead Conversion Detail contains no rows."
    elif selected_range and not filtered_rows:
        source_status = "empty"
        source_message = "HR conversion workbook is present but no conversion rows matched the selected lead intake date range."
    elif missing_tabs:
        source_status = "partial"
        source_message = f"HR conversion workbook is missing optional tab(s): {', '.join(missing_tabs)}."

    current_summary = _conversion_metrics(filtered_rows)
    return {
        "workbook_name": path.name,
        "source_status": source_status,
        "source_message": source_message,
        "source_system": CONVERSION_SOURCE_SYSTEM,
        "source_authority": CONVERSION_SOURCE_AUTHORITY,
        "projection_mode": "read_only",
        "pii_suppressed": True,
        "conversion_rule": _conversion_rule(workbook),
        "date_range": selected_range.as_dict() if selected_range else None,
        "tabs": _conversion_tab_counts(workbook, present_tabs),
        "missing_tabs": missing_tabs,
        "summary": {
            **current_summary,
            "unfiltered_eligible_leads": len(detail_rows),
        },
        "trend_summary": _conversion_trend_summary(detail_rows, selected_range),
        "by_source": _conversion_group_rows(filtered_rows, CONVERSION_SOURCE_ALIASES, "source_bucket"),
        "by_sla": _conversion_group_rows(filtered_rows, CONVERSION_SLA_ALIASES, "sla_result"),
        "trend": _conversion_trend_rows(filtered_rows),
    }


def _empty_conversion_funnel(
    *,
    workbook_name: str | None = None,
    source_status: str,
    source_message: str,
    present_tabs: list[str] | None = None,
    missing_tabs: list[str] | None = None,
    selected_range: DashboardDateRange | None = None,
) -> dict[str, Any]:
    return {
        "workbook_name": workbook_name,
        "source_status": source_status,
        "source_message": source_message,
        "source_system": CONVERSION_SOURCE_SYSTEM,
        "source_authority": CONVERSION_SOURCE_AUTHORITY,
        "projection_mode": "read_only",
        "pii_suppressed": True,
        "conversion_rule": "Exact normalized lead name match against the Xcelerator driver list.",
        "date_range": selected_range.as_dict() if selected_range else None,
        "tabs": [
            {"sheet": sheet, "row_count": 0, "status": "present"}
            for sheet in (present_tabs or [])
        ],
        "missing_tabs": list(missing_tabs or CONVERSION_EXPECTED_TABS),
        "summary": {
            **_conversion_metrics([]),
            "unfiltered_eligible_leads": 0,
        },
        "trend_summary": None,
        "by_source": [],
        "by_sla": [],
        "trend": [],
    }


def _conversion_metrics(records: list[dict[str, Any]]) -> dict[str, Any]:
    eligible = len(records)
    converted = sum(1 for record in records if _conversion_yes(record, CONVERSION_CONVERTED_ALIASES))
    still_driving = sum(1 for record in records if _conversion_yes(record, CONVERSION_STILL_DRIVING_ALIASES))
    phone_match = sum(1 for record in records if _conversion_yes(record, CONVERSION_PHONE_MATCH_ALIASES))
    return {
        "eligible_leads": eligible,
        "converted_leads": converted,
        "not_converted_leads": max(eligible - converted, 0),
        "conversion_rate": round(converted / eligible, 4) if eligible else None,
        "still_driving_count": still_driving,
        "still_driving_rate": round(still_driving / eligible, 4) if eligible else None,
        "phone_match_count": phone_match,
    }


def _conversion_trend_summary(
    records: list[dict[str, Any]],
    selected_range: DashboardDateRange | None,
) -> dict[str, Any] | None:
    if not selected_range:
        return None
    previous_range = selected_range.previous()
    current = _conversion_metrics(
        _filter_records_by_date_range(records, selected_range, CONVERSION_INTAKE_DATE_FIELDS)
    )
    previous = _conversion_metrics(
        _filter_records_by_date_range(records, previous_range, CONVERSION_INTAKE_DATE_FIELDS)
    )
    current_rate = current["conversion_rate"]
    previous_rate = previous["conversion_rate"]
    return {
        "current": current,
        "previous": previous,
        "eligible_change_pct": pct_change(current["eligible_leads"], previous["eligible_leads"]),
        "converted_change_pct": pct_change(current["converted_leads"], previous["converted_leads"]),
        "conversion_rate_change_points": (
            round((current_rate - previous_rate) * 100, 1)
            if current_rate is not None and previous_rate is not None
            else None
        ),
    }


def _conversion_group_rows(
    records: list[dict[str, Any]],
    aliases: tuple[str, ...],
    label_key: str,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        label = _text(_value_for(record, aliases), "Unknown")
        grouped[label].append(record)

    rows: list[dict[str, Any]] = []
    for label, group in grouped.items():
        rows.append({label_key: label, **_conversion_metrics(group)})
    return sorted(
        rows,
        key=lambda row: (-int(row["eligible_leads"]), str(row[label_key])),
    )


def _conversion_trend_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        parsed = _parse_datetime(_value_for(record, CONVERSION_INTAKE_DATE_FIELDS))
        if parsed is None:
            continue
        grouped[parsed.date().isoformat()].append(record)
    return [
        {"date": day, **_conversion_metrics(group)}
        for day, group in sorted(grouped.items(), key=lambda item: item[0])
    ]


def _conversion_tab_counts(workbook: Workbook, present_tabs: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    required_by_sheet = {
        "Summary": ("Metric", "Value"),
        "Lead Conversion Detail": ("Lead Created At", "Source Systems", "Converted By Rule"),
        "Converted Leads": ("Lead Created At", "Converted By Rule"),
        "Not Converted Leads": ("Lead Created At", "Converted By Rule"),
        "Conversion by Source": ("Source Systems", "Eligible Leads", "Converted by Name"),
        "Rules and Issues": ("Item", "Detail"),
    }
    for sheet in present_tabs:
        records = _records_for_sheet(
            workbook,
            sheet,
            required_headers=required_by_sheet.get(sheet, ("Lead Created At", "Converted By Rule")),
        )
        rows.append({"sheet": sheet, "row_count": len(records), "status": "present"})
    return rows


def _conversion_rule(workbook: Workbook) -> str:
    rows = _records_for_sheet(workbook, "Rules and Issues", required_headers=("Item", "Detail"))
    for row in rows:
        item = _normalize_header(_text(row.get("Item")))
        if item == "conversionruleapplied":
            return _text(row.get("Detail"), "Exact normalized lead name match against the Xcelerator driver list.")
    return "Exact normalized lead name match against the Xcelerator driver list."


def _conversion_yes(record: dict[str, Any], aliases: tuple[str, ...]) -> bool:
    return _boolish(_value_for(record, aliases))


def _conversion_count(conversion_funnel: dict[str, Any] | None, key: str) -> int:
    if not conversion_funnel:
        return 0
    summary = conversion_funnel.get("summary")
    if not isinstance(summary, dict):
        return 0
    return int(summary.get(key) or 0)


def _exception_queue_rows(
    lead_rows: list[dict[str, Any]],
    exception_records_by_tab: dict[str, list[dict[str, Any]]],
    *,
    as_of: datetime,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    for sheet in EXCEPTION_TABS:
        spec = EXCEPTION_TAB_SPECS[sheet]
        for record in exception_records_by_tab.get(sheet, []):
            row = _exception_queue_row(
                record,
                source_sheet=sheet,
                category=str(spec["category"]),
                reason=str(spec["reason"]),
                severity=str(spec["severity"]),
                as_of=as_of,
            )
            _append_unique_exception(rows, seen, row)

    for record in lead_rows:
        for spec in _lead_level_exception_specs(record):
            row = _exception_queue_row(
                record,
                source_sheet="Lead Level KPI",
                category=str(spec["category"]),
                reason=str(spec["reason"]),
                severity=str(spec["severity"]),
                as_of=as_of,
            )
            _append_unique_exception(rows, seen, row)

    severity_order = {"critical": 0, "warning": 1}
    return sorted(
        rows,
        key=lambda row: (
            severity_order.get(str(row["severity"]), 9),
            -float(row["age_hours"] or 0),
            str(row["lead_created_date"] or ""),
            str(row["category"]),
        ),
    )


def _append_unique_exception(
    rows: list[dict[str, Any]],
    seen: set[tuple[str, str]],
    row: dict[str, Any],
) -> None:
    key = (str(row["category"]), str(row["lead_ref"]))
    if key in seen:
        return
    seen.add(key)
    rows.append(row)


def _lead_level_exception_specs(record: dict[str, Any]) -> list[dict[str, str]]:
    specs: list[dict[str, str]] = []
    first_bucket = _text(_value_for(record, FIRST_OUTREACH_BUCKET_ALIASES))
    discussion_bucket = _text(_value_for(record, REAL_DISCUSSION_BUCKET_ALIASES))
    first_key = _normalize_header(first_bucket)
    discussion_key = _normalize_header(discussion_bucket)
    if first_key == "nooutboundfound":
        specs.append(EXCEPTION_TAB_SPECS["Failed No Outbound"])
    if first_key == "over72hfailed":
        specs.append(EXCEPTION_TAB_SPECS["Failed Over 72h"])
    if discussion_key == "no1mindiscussionfound":
        specs.append(EXCEPTION_TAB_SPECS["No Real Discussion"])
    return specs


def _exception_queue_row(
    record: dict[str, Any],
    *,
    source_sheet: str,
    category: str,
    reason: str,
    severity: str,
    as_of: datetime,
) -> dict[str, Any]:
    lead_key = _lead_identity_key(record)
    lead_created_at = _parse_datetime(_value_for(record, LEAD_CREATED_ALIASES))
    first_bucket = _text(_value_for(record, FIRST_OUTREACH_BUCKET_ALIASES), "Unknown")
    discussion_bucket = _text(_value_for(record, REAL_DISCUSSION_BUCKET_ALIASES), "Unknown")
    return {
        "exception_id": _hash_key((category, lead_key)),
        "lead_ref": f"lead-ref-{lead_key[:4]}-{lead_key[4:8]}",
        "masked_contact": "PII suppressed",
        "category": category,
        "reason": reason,
        "severity": severity,
        "source_sheet": source_sheet,
        "source_system": SOURCE_SYSTEM,
        "source_authority": SOURCE_AUTHORITY,
        "lead_created_date": lead_created_at.date().isoformat() if lead_created_at else None,
        "status": _text(_value_for(record, STATUS_ALIASES), "Unknown"),
        "first_outreach_bucket": first_bucket,
        "real_discussion_bucket": discussion_bucket,
        "age_hours": _exception_age_hours(record, category=category, lead_created_at=lead_created_at, as_of=as_of),
        "pii_suppressed": True,
        "projection_mode": "read_only",
    }


def _exception_age_hours(
    record: dict[str, Any],
    *,
    category: str,
    lead_created_at: datetime | None,
    as_of: datetime,
) -> float | None:
    aliases = DISCUSSION_HOUR_ALIASES if category == "No Real Discussion" else OUTREACH_HOUR_ALIASES
    explicit_hours = _number(_value_for(record, aliases))
    if explicit_hours is not None:
        return round(explicit_hours, 2)
    if lead_created_at is None:
        return None
    return round(max((as_of - lead_created_at).total_seconds() / 3600, 0), 2)


def _lead_identity_key(record: dict[str, Any]) -> str:
    parts = [_text(_value_for(record, (alias,))) for alias in IDENTITY_ALIASES]
    parts.extend(_text(_value_for(record, (alias,))) for alias in (*LEAD_CREATED_ALIASES, *STATUS_ALIASES))
    stable_parts = [part for part in parts if part]
    if not stable_parts:
        stable_parts = [_text(value) for value in record.values() if _text(value)][:8]
    return _hash_key(tuple(stable_parts or ["unknown-lead"]))


def _value_for(record: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    normalized = {_normalize_header(alias) for alias in aliases}
    for key, value in record.items():
        if _normalize_header(str(key)) in normalized:
            return value
    return None


def _hash_key(parts: tuple[str, ...]) -> str:
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]


def _exception_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts = Counter(str(row["category"]) for row in rows)
    return [{"category": category, "count": counts[category]} for category in sorted(counts)]


def _bucket_counts(counter: Counter[str], order: tuple[str, ...]) -> list[dict[str, Any]]:
    ordered_keys = [key for key in order if key in counter]
    ordered_keys.extend(sorted(key for key in counter if key not in ordered_keys))
    return [{"bucket": key, "count": int(counter[key])} for key in ordered_keys]


def _build_hard_target_results(
    hard_targets: dict[str, dict[str, Any]],
    actuals: dict[str, int | float | None],
    evidence_available: dict[str, bool],
) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}
    for key, spec in hard_targets.items():
        actual = actuals.get(key)
        can_evaluate = evidence_available.get(key, actual is not None) and actual is not None
        status = (
            "awaiting_feed"
            if not can_evaluate
            else "healthy"
            if _target_met(float(actual), str(spec["operator"]), float(spec["target"]))
            else "warning"
        )
        results[key] = {
            "key": key,
            "label": spec["label"],
            "actual": actual,
            "target": spec["target"],
            "operator": spec["operator"],
            "unit": spec["unit"],
            "cadence": spec["cadence"],
            "display_target": spec["display_target"],
            "status": status,
        }
    return results


def _target_met(actual: float, operator: str, target: float) -> bool:
    if operator == ">=":
        return actual >= target
    if operator == "<=":
        return actual <= target
    return actual == target


def _normalize_header(value: str) -> str:
    return "".join(ch for ch in value.casefold() if ch.isalnum())


def _ensure_aware(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return _ensure_aware(value)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=timezone.utc)
    if isinstance(value, (int, float)) and float(value) > 20000:
        return datetime(1899, 12, 30, tzinfo=timezone.utc) + timedelta(days=float(value))
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in (
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%y %I:%M %p",
        "%m/%d/%Y %H:%M",
        "%m/%d/%y %H:%M",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        return _ensure_aware(datetime.fromisoformat(raw.replace("Z", "+00:00")))
    except ValueError:
        return None


def _same_day(value: Any, as_of: datetime) -> bool:
    parsed = _parse_datetime(value)
    return bool(parsed and parsed.date() == as_of.date())


def _record_matches_date_range(
    record: dict[str, Any],
    selected_range: DashboardDateRange | None,
    fields: tuple[str, ...],
) -> bool:
    if not selected_range:
        return True
    return any(
        selected_range.contains_datetime(_parse_datetime(record.get(field)))
        for field in fields
    )


def _filter_records_by_date_range(
    records: list[dict[str, Any]],
    selected_range: DashboardDateRange | None,
    fields: tuple[str, ...],
) -> list[dict[str, Any]]:
    if not selected_range:
        return records
    return [
        record
        for record in records
        if _record_matches_date_range(record, selected_range, fields)
    ]


def _period_metrics(
    leads: list[dict[str, Any]],
    call_attempts: list[dict[str, Any]],
    selected_range: DashboardDateRange | None,
) -> dict[str, int]:
    return {
        "new_leads": _record_count_for_range(leads, selected_range, LEAD_INTAKE_DATE_FIELDS),
        "new_applicants": len(leads),
        "interviews_scheduled": _record_count_for_range(
            leads,
            selected_range,
            LEAD_OUTCOME_DATE_FIELDS[:2],
        ),
        "new_hires": _record_count_for_range(leads, selected_range, LEAD_OUTCOME_DATE_FIELDS[2:]),
        "follow_ups": len(call_attempts),
    }


def _record_count_for_range(
    records: list[dict[str, Any]],
    selected_range: DashboardDateRange | None,
    fields: tuple[str, ...],
) -> int:
    if not selected_range:
        return len(records)
    return sum(1 for record in records if _record_matches_date_range(record, selected_range, fields))


def _trend_comparison(
    leads: list[dict[str, Any]],
    call_attempts: list[dict[str, Any]],
    selected_range: DashboardDateRange | None,
) -> dict[str, Any] | None:
    if not selected_range:
        return None
    previous_range = selected_range.previous()
    current = _period_metrics(
        _filter_records_by_date_range(leads, selected_range, LEAD_INTAKE_DATE_FIELDS),
        _filter_records_by_date_range(call_attempts, selected_range, CALL_ATTEMPT_DATE_FIELDS),
        selected_range,
    )
    previous = _period_metrics(
        _filter_records_by_date_range(leads, previous_range, LEAD_INTAKE_DATE_FIELDS),
        _filter_records_by_date_range(call_attempts, previous_range, CALL_ATTEMPT_DATE_FIELDS),
        previous_range,
    )
    return {
        "current": current,
        "previous": previous,
        "lead_volume_change_pct": pct_change(current["new_leads"], previous["new_leads"]),
        "hire_volume_change_pct": pct_change(current["new_hires"], previous["new_hires"]),
        "follow_up_change_pct": pct_change(current["follow_ups"], previous["follow_ups"]),
    }


def _text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    try:
        return round(float(str(value).strip().replace("%", "")), 4)
    except ValueError:
        return None


def _int(value: Any) -> int:
    number = _number(value)
    return int(number or 0)


def _boolish(value: Any) -> bool:
    normalized = _normalize_header(_text(value))
    return normalized in {"yes", "y", "true", "1", "used", "mapped"}


def _mean(values: list[float] | None) -> float:
    values = values or []
    return round(sum(values) / len(values), 2) if values else 0.0


def _max_or_zero(values: list[float] | None) -> float:
    values = values or []
    return round(max(values), 2) if values else 0.0
