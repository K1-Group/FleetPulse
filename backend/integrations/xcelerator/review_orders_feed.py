"""Read-only ReviewOrders feed loader for Xcelerator lane stability scoring."""

from __future__ import annotations

import csv
import json
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import httpx


@dataclass(frozen=True)
class ReviewOrdersFeedConfig:
    """Configuration for a read-only Xcelerator ReviewOrders projection."""

    url: str = ""
    path: str = ""
    api_key: str = ""
    api_key_header: str = "X-FleetPulse-Lane-Stability-Key"
    timeout_seconds: float = 30.0

    @classmethod
    def from_env(cls, prefix: str = "FLEETPULSE_LANE_STABILITY") -> "ReviewOrdersFeedConfig":
        timeout_raw = os.getenv(f"{prefix}_TIMEOUT_SECONDS", "30")
        try:
            timeout_seconds = float(timeout_raw)
        except ValueError:
            timeout_seconds = 30.0
        return cls(
            url=os.getenv(f"{prefix}_ORDER_FEED_URL", "").strip(),
            path=os.getenv(f"{prefix}_ORDER_FEED_PATH", "").strip(),
            api_key=os.getenv(f"{prefix}_ORDER_FEED_API_KEY", "").strip(),
            api_key_header=os.getenv(
                f"{prefix}_ORDER_FEED_API_KEY_HEADER",
                "X-FleetPulse-Lane-Stability-Key",
            ).strip()
            or "X-FleetPulse-Lane-Stability-Key",
            timeout_seconds=timeout_seconds,
        )

    @property
    def configured(self) -> bool:
        return bool(self.url or self.path)


def _coerce_rows(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if not isinstance(payload, dict):
        return []

    for key in ("rows", "value", "orders", "data", "items"):
        value = payload.get(key)
        if isinstance(value, list):
            return [row for row in value if isinstance(row, dict)]

    tables = payload.get("tables")
    if isinstance(tables, dict):
        for key in ("Orders", "orders", "ReviewOrders", "review_orders"):
            value = tables.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
        for value in tables.values():
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]

    return []


def _load_json_or_csv(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.casefold()
    if suffix == ".json":
        return _coerce_rows(json.loads(path.read_text(encoding="utf-8")))
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as exc:
            raise RuntimeError("openpyxl_required_for_xlsx_review_orders_feed") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        records: list[dict[str, Any]] = []
        for raw_row in rows[1:]:
            record = {
                headers[index]: value
                for index, value in enumerate(raw_row)
                if index < len(headers) and headers[index]
            }
            if any(value not in {None, ""} for value in record.values()):
                records.append(record)
        return records
    raise RuntimeError(f"unsupported_review_orders_feed_file:{suffix or 'unknown'}")


def load_review_orders_rows(config: ReviewOrdersFeedConfig) -> list[dict[str, Any]]:
    """Load raw Xcelerator ReviewOrders rows from a configured URL or file path."""

    if config.url:
        headers = {"Accept": "application/json"}
        if config.api_key:
            headers[config.api_key_header] = config.api_key
        with httpx.Client(timeout=config.timeout_seconds) as client:
            response = client.get(config.url, headers=headers)
        response.raise_for_status()
        return _coerce_rows(response.json())

    if config.path:
        return _load_json_or_csv(Path(config.path))

    return []

