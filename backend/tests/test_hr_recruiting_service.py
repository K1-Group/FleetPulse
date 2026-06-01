"""Tests for HR recruiting worklist calculations."""

from __future__ import annotations

import json
import asyncio
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

from services.hr_recruiting_service import (  # noqa: E402
    HrRecruitingConfig,
    build_hr_recruiting_dataset,
    get_hr_recruiting_dataset,
    import_hr_recruiting_snapshot,
)


def _config() -> HrRecruitingConfig:
    return HrRecruitingConfig(
        table_id="01KR00WV4YHCB6BMYDE1EG7HEM",
        source="zapier_table",
        sla_hours=(24, 48, 72),
    )


def test_hr_recruiting_dataset_calculates_age_stale_and_process_time() -> None:
    now = datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc)
    rows = [
        {
            "source_email_id": "outlook-message-1",
            "applicant": "Private Applicant One",
            "worklist": "Recruiter Review",
            "status": "Assigned",
            "first_assigned_at": "2026-05-14T10:00:00Z",
            "current_worklist_entered_at": "2026-05-15T00:00:00Z",
            "first_contacted_at": "2026-05-14T12:00:00Z",
            "qualified": "yes",
            "phone": "555-0100",
            "email": "private.one@example.com",
            "ssn": "123-45-6789",
        },
        {
            "source_email_id": "outlook-message-1",
            "applicant": "Private Applicant One",
            "worklist": "Recruiter Review",
            "status": "Assigned",
            "first_assigned_at": "2026-05-14T10:00:00Z",
            "current_worklist_entered_at": "2026-05-15T00:00:00Z",
            "first_contacted_at": "2026-05-14T12:00:00Z",
            "qualified": "yes",
        },
        {
            "applicant": "Private Applicant Two",
            "worklist": "Safety Review",
            "status": "Qualified",
            "first_assigned_at": "2026-05-12T10:00:00Z",
            "current_worklist_entered_at": "2026-05-12T12:00:00Z",
        },
        {
            "source_email_id": "outlook-message-3",
            "applicant": "Private Applicant Three",
            "worklist": "Background Check",
            "status": "Completed",
            "first_assigned_at": "2026-05-15T08:00:00Z",
            "current_worklist_entered_at": "2026-05-15T08:00:00Z",
            "first_contacted_at": "2026-05-15T09:00:00Z",
            "completed_at": "2026-05-15T11:00:00Z",
            "hired_at": "2026-05-15T11:00:00Z",
            "orientation_scheduled_at": "2026-05-15T10:00:00Z",
            "orientation_showed": "yes",
        },
    ]

    dataset = build_hr_recruiting_dataset(rows, now=now, config=_config())

    assert dataset["projection_mode"] == "read_only"
    assert dataset["pii_suppressed"] is True
    assert dataset["row_counts"]["source_rows"] == 4
    assert dataset["row_counts"]["deduped_leads"] == 3
    assert dataset["summary"] == {
        "active_leads": 2,
        "new_leads_today": 1,
        "avg_process_age_hours": 50.0,
        "stale_leads": 1,
        "completed_today": 1,
        "new_hires_7d": 1,
        "active_qualified_pipeline": 2,
        "first_touch_24h_pct": 0.6667,
        "first_touch_eligible_count": 3,
        "first_touch_within_24h_count": 2,
        "stale_untouched_48h": 1,
        "orientation_scheduled_count": 1,
        "orientation_show_count": 1,
        "orientation_show_rate": 1.0,
    }
    assert dataset["hard_target_status"] == "warning"
    assert dataset["hard_targets"]["active_qualified_pipeline"]["target"] == 10
    assert dataset["hard_targets"]["active_qualified_pipeline"]["status"] == "warning"
    assert dataset["hard_targets"]["new_hires_7d"]["display_target"] == ">= 5/week"
    assert dataset["hard_targets"]["first_touch_24h_pct"]["display_target"] == ">= 95% within 24h"
    assert dataset["hard_targets"]["orientation_show_rate"]["status"] == "healthy"
    assert "stale_untouched_48h" in dataset["hard_target_misses"]
    assert dataset["team_members"] == [
        {
            "name": "Jordan",
            "configured": True,
            "status": "configured",
            "evidence_sources": [],
            "lead_count": 0,
            "first_outreach_leads": 0,
            "real_discussion_leads": 0,
            "within_24h": 0,
            "recovered_24_48h": 0,
            "late_48_72h": 0,
            "failed_over_72h": 0,
            "total_outbound_attempts": 0,
            "within_24h_rate": None,
        }
    ]

    by_worklist = {row["worklist"]: row for row in dataset["by_worklist"]}
    assert by_worklist["Recruiter Review"]["avg_age_hours"] == 12.0
    assert by_worklist["Recruiter Review"]["stale_24h"] == 0
    assert by_worklist["Safety Review"]["max_age_hours"] == 72.0
    assert by_worklist["Safety Review"]["stale_24h"] == 1
    assert by_worklist["Safety Review"]["stale_48h"] == 1
    assert by_worklist["Safety Review"]["stale_72h"] == 1

    daily = {(row["date"], row["worklist"]): row for row in dataset["daily"]}
    assert daily[("2026-05-15", "Background Check")]["completed_leads"] == 1
    assert daily[("2026-05-15", "Background Check")]["avg_process_time_hours"] == 3.0

    trend = {row["date"]: row for row in dataset["trend"]}
    assert trend["2026-05-12"]["active_leads"] == 1
    assert trend["2026-05-12"]["stale_leads"] == 1
    assert trend["2026-05-14"]["avg_age_hours"] == 26.0

    serialized = json.dumps(dataset)
    assert "private.one@example.com" not in serialized
    assert "555-0100" not in serialized
    assert "123-45-6789" not in serialized
    assert "Private Applicant" not in serialized


def test_fallback_dedupe_hash_uses_applicant_worklist_and_assignment_time() -> None:
    now = datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc)
    rows = [
        {
            "applicant": "Private Applicant",
            "worklist": "Recruiter Review",
            "status": "Assigned",
            "first_assigned_at": "05/15/2026 08:00 AM",
        },
        {
            "applicant": "Private Applicant",
            "worklist": "Recruiter Review",
            "status": "Assigned",
            "first_assigned_at": "05/15/2026 08:00 AM",
        },
    ]

    dataset = build_hr_recruiting_dataset(rows, now=now, config=_config())

    assert dataset["row_counts"]["source_rows"] == 2
    assert dataset["row_counts"]["deduped_leads"] == 1
    assert dataset["summary"]["active_leads"] == 1
    assert dataset["summary"]["new_leads_today"] == 1


def test_hr_recruiting_dataset_filters_custom_date_range_and_compares_previous() -> None:
    now = datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc)
    rows = [
        {
            "applicant": "Private Current Applicant",
            "worklist": "Recruiter Review",
            "status": "Assigned",
            "first_assigned_at": "2026-05-10T08:00:00Z",
            "first_contacted_at": "2026-05-10T09:00:00Z",
        },
        {
            "applicant": "Private Previous Applicant",
            "worklist": "Recruiter Review",
            "status": "Hired",
            "first_assigned_at": "2026-05-03T08:00:00Z",
            "first_contacted_at": "2026-05-03T09:00:00Z",
            "hired_at": "2026-05-03T12:00:00Z",
        },
    ]

    dataset = build_hr_recruiting_dataset(
        rows,
        now=now,
        config=_config(),
        start_date="2026-05-10",
        end_date="2026-05-10",
    )

    assert dataset["date_range"]["start"] == "2026-05-10"
    assert dataset["row_counts"]["deduped_leads"] == 1
    assert dataset["row_counts"]["unfiltered_deduped_leads"] == 2
    assert dataset["period_metrics"] == {
        "new_leads": 1,
        "new_applicants": 1,
        "interviews_scheduled": 0,
        "new_hires": 0,
        "follow_ups": 1,
    }
    assert dataset["trend_comparison"]["previous"]["new_leads"] == 0
    assert dataset["trend_comparison"]["lead_volume_change_pct"] == 100.0


def test_hr_recruiting_date_windows_cover_today_last_7_and_last_30_days() -> None:
    now = datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc)
    rows = [
        {
            "applicant": "Private Today",
            "worklist": "Recruiter Review",
            "status": "Assigned",
            "first_assigned_at": "2026-05-15T08:00:00Z",
        },
        {
            "applicant": "Private Week",
            "worklist": "Recruiter Review",
            "status": "Assigned",
            "first_assigned_at": "2026-05-10T08:00:00Z",
        },
        {
            "applicant": "Private Month",
            "worklist": "Recruiter Review",
            "status": "Assigned",
            "first_assigned_at": "2026-04-20T08:00:00Z",
        },
    ]

    today = build_hr_recruiting_dataset(
        rows,
        now=now,
        config=_config(),
        start_date="2026-05-15",
        end_date="2026-05-15",
    )
    last_7 = build_hr_recruiting_dataset(
        rows,
        now=now,
        config=_config(),
        start_date="2026-05-09",
        end_date="2026-05-15",
    )
    last_30 = build_hr_recruiting_dataset(
        rows,
        now=now,
        config=_config(),
        start_date="2026-04-16",
        end_date="2026-05-15",
    )

    assert today["period_metrics"]["new_leads"] == 1
    assert last_7["period_metrics"]["new_leads"] == 2
    assert last_30["period_metrics"]["new_leads"] == 3


def test_empty_hr_recruiting_dataset_is_explicit_and_safe() -> None:
    now = datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc)

    dataset = build_hr_recruiting_dataset(
        [],
        now=now,
        config=_config(),
        source_status="snapshot_not_configured",
        source_message="Configure HR_RECRUITING_SNAPSHOT_URL.",
    )

    assert dataset["source_status"] == "snapshot_not_configured"
    assert dataset["summary"]["active_leads"] == 0
    assert dataset["hard_target_status"] == "awaiting_feed"
    assert dataset["hard_targets"]["active_qualified_pipeline"]["status"] == "awaiting_feed"
    assert dataset["by_worklist"] == []
    assert dataset["daily"] == []
    assert dataset["status_counts"] == []
    assert dataset["trend"] == []
    assert dataset["source_message"] == "Configure HR_RECRUITING_SNAPSHOT_URL."


def test_imported_hr_recruiting_state_feeds_dataset_without_pii(tmp_path) -> None:
    state_path = tmp_path / "hr-recruiting.json"
    result = import_hr_recruiting_snapshot(
        json.dumps(
            {
                "rows": [
                    {
                        "source_email_id": "outlook-message-1",
                        "applicant": "Private Applicant",
                        "worklist": "Recruiter Review",
                        "status": "Assigned",
                        "first_assigned_at": "2026-05-14T10:00:00Z",
                        "current_worklist_entered_at": "2026-05-15T00:00:00Z",
                        "email": "private@example.com",
                    }
                ]
            }
        ),
        filename="hr-recruiting.json",
        path=state_path,
    )

    assert result["status"] == "ok"
    assert result["row_count"] == 1
    dataset = asyncio.run(
        get_hr_recruiting_dataset(
            now=datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc),
            config=HrRecruitingConfig(snapshot_path=str(state_path), source="zapier_table"),
        )
    )

    assert dataset["source_status"] == "ok"
    assert dataset["summary"]["active_leads"] == 1
    serialized = json.dumps(dataset)
    assert "private@example.com" not in serialized
    assert "Private Applicant" not in serialized


def test_hr_recruiting_import_dry_run_does_not_write(tmp_path) -> None:
    state_path = tmp_path / "hr-recruiting.json"
    result = import_hr_recruiting_snapshot(
        "source_email_id,applicant,worklist,status,first_assigned_at\nmsg-1,Private Applicant,Recruiter Review,Assigned,2026-05-14",
        filename="hr.csv",
        dry_run=True,
        path=state_path,
    )

    assert result["status"] == "ok"
    assert not state_path.exists()


def _write_hr_kpi_workbook(path: Path, *, include_modified_backlog: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Level KPI"
    ws.append(["HR Lead KPI Recheck"])
    ws.append(
        [
            "Lead Name",
            "Phone",
            "Email",
            "Lead Created At",
            "App Status",
            "First Outreach KPI Bucket",
            "Hours to First Outreach",
            "Real Discussion KPI Bucket",
            "Hours to First Real Discussion",
            "Modified At",
        ]
    )
    ws.append(["Private Applicant One", "555-0100", "one@example.com", "2026-05-14 08:00", "Not Qualified", "Within 24h", 2, "Within 24h", 3, None])
    ws.append(["Private Applicant Two", "555-0101", "two@example.com", "2026-05-14 09:00", "New", "No outbound found", None, "No 1min+ discussion found", None, None])
    ws.append(["Private Applicant Three", "555-0102", "three@example.com", "2026-05-13 09:00", "No Response", "Over 72h failed", 80, "Over 72h failed", 82, None])
    if include_modified_backlog:
        ws.append(["Private Backlog Applicant", "555-0199", "backlog@example.com", "2026-03-20 08:00", "Not Qualified", "No outbound found", None, "No 1min+ discussion found", None, "2026-05-14 11:00"])

    ws = wb.create_sheet("Call Attempts Detail")
    ws.append(["Detail"])
    ws.append(["Call Date/Time", "HR Member", "Duration Seconds", "Real Discussion 1min+"])
    ws.append(["2026-05-14 10:00", "Recruiter A", 180, "Yes"])
    ws.append(["2026-05-16 10:00", "Recruiter A", 20, "No"])

    ws = wb.create_sheet("KPI By First Outreach")
    ws.append(["Summary"])
    ws.append(["HR Member", "Leads_First_Outreach", "Within_24h", "Recovered_24_48", "Late_48_72", "Failed_Over_72", "Avg_Hours_To_First_Outreach", "Median_Hours_To_First_Outreach", "Total_Outbound_Attempts", "Within_24h_Rate"])
    ws.append(["Recruiter A", 2, 1, 0, 0, 1, 41, 41, 2, 0.5])

    ws = wb.create_sheet("KPI By Real Discussion")
    ws.append(["Summary"])
    ws.append(["HR Member", "Leads_First_Real_Discussion", "Within_24h", "Recovered_24_48", "Late_48_72", "Failed_Over_72", "Avg_Hours_To_First_Real", "Median_Hours_To_First_Real", "Real_Within_24_Rate"])
    ws.append(["Recruiter A", 2, 1, 0, 0, 1, 42.5, 42.5, 0.5])

    exception_headers = [
        "Lead Name",
        "Phone",
        "Email",
        "Lead Created At",
        "App Status",
        "First Outreach KPI Bucket",
        "Hours to First Outreach",
        "Real Discussion KPI Bucket",
        "Hours to First Real Discussion",
        "Modified At",
    ]
    for sheet in ("Failed No Outbound", "Recovered 24-48h", "Failed Over 72h", "No Real Discussion"):
        ws = wb.create_sheet(sheet)
        ws.append(["Subset"])
        ws.append(exception_headers)
        if sheet == "Failed No Outbound":
            ws.append(["Private Applicant Two", "555-0101", "two@example.com", "2026-05-14 09:00", "New", "No outbound found", None, "No 1min+ discussion found", None, None])
            if include_modified_backlog:
                ws.append(["Private Backlog Applicant", "555-0199", "backlog@example.com", "2026-03-20 08:00", "Not Qualified", "No outbound found", None, "No 1min+ discussion found", None, "2026-05-14 11:00"])
        if sheet == "Failed Over 72h":
            ws.append(["Private Applicant Three", "555-0102", "three@example.com", "2026-05-13 09:00", "No Response", "Over 72h failed", 80, "Over 72h failed", 82, None])
        if sheet == "No Real Discussion":
            ws.append(["Private Applicant Two", "555-0101", "two@example.com", "2026-05-14 09:00", "New", "No outbound found", None, "No 1min+ discussion found", None, None])

    ws = wb.create_sheet("Source Log QA")
    ws.append(["File", "Rows", "Columns", "Used for Mapping", "Reason / Notes", "First Columns"])
    ws.append(["Detail_HR.csv", 100, 9, "Yes", "Phone-level call detail used for lead matching", "Date/Time, Extension"])
    ws.append(["Report_Export.csv", 12, 23, "No", "Tenstreet status summary; not a call log", "Source, Referrer"])
    wb.save(path)


def _write_hr_conversion_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([])
    ws.append(["Lead Name to Xcelerator Driver Conversion"])
    ws.append(["Rule used: exact normalized lead name found in Xcelerator driver list = converted/processed."])
    ws.append([])
    ws.append(["Metric", "Value", "Definition"])
    ws.append(["Eligible unique leads reviewed", 3, "From source-backed HR lead files"])
    ws.append(["Converted by exact name match", 2, "66.7%"])
    ws.append(["Not converted by exact name match", 1, "33.3%"])

    detail_headers = [
        "Lead Name",
        "Lead Phone",
        "Lead Created At",
        "Source Systems",
        "First Outbound Member",
        "SLA Result",
        "Converted By Rule",
        "Conversion Rule Evidence",
        "Driver No",
        "Driver Name",
        "Driver Status",
        "Driver Start Date",
        "Driver Depart Date",
        "Still Driving Evidence",
        "Phone Also Matches Driver",
        "Driver Phone",
        "Driver Mobile",
        "Name Match Count",
        "Phone Match Count",
        "Application Date",
        "App Status",
        "App Worklist",
    ]
    detail_rows = [
        [
            "Private Converted One",
            "555-0100",
            "2026-05-14 08:00",
            "Facebook Leads File",
            "Recruiter A",
            "Called within 24h",
            "Yes",
            "Exact normalized lead name exists in Xcelerator driver list",
            "DX-1",
            "Private Converted One",
            "A",
            "2026-05-20 00:00",
            None,
            "Yes",
            "Yes",
            "555-0100",
            "555-0100",
            1,
            1,
            "2026-05-14 10:00",
            "Hired",
            "DFW, TX",
        ],
        [
            "Private Not Converted",
            "555-0101",
            "2026-05-14 09:00",
            "Microsoft Teams - Onboarding Drivers",
            "Recruiter B",
            "No outbound call found",
            "No",
            "Lead normalized name not found in Xcelerator driver list",
            None,
            None,
            None,
            None,
            None,
            "No",
            "No",
            None,
            None,
            0,
            0,
            "2026-05-14 11:00",
            "Not Qualified",
            "DFW, TX",
        ],
        [
            "Private Previous Converted",
            "555-0102",
            "2026-05-13 09:00",
            "Facebook Leads File",
            "Recruiter A",
            "Called after 24h",
            "Yes",
            "Exact normalized lead name exists in Xcelerator driver list",
            "DX-2",
            "Private Previous Converted",
            "I",
            "2026-05-18 00:00",
            None,
            "No",
            "No",
            "555-0102",
            None,
            1,
            0,
            "2026-05-13 11:00",
            "Hired",
            "DFW, TX",
        ],
    ]

    for sheet in ("Lead Conversion Detail", "Converted Leads", "Not Converted Leads"):
        ws = wb.create_sheet(sheet)
        ws.append([])
        ws.append([sheet])
        ws.append(["Exact-name match results against Xcelerator-All-drivers.xlsx."])
        ws.append([])
        ws.append(detail_headers)
        for row in detail_rows:
            converted = row[6] == "Yes"
            if sheet == "Converted Leads" and not converted:
                continue
            if sheet == "Not Converted Leads" and converted:
                continue
            ws.append(row)

    ws = wb.create_sheet("Conversion by Source")
    ws.append([])
    ws.append(["Conversion by Source"])
    ws.append(["Source bucket conversion using exact-name rule."])
    ws.append([])
    ws.append(["Source Systems", "Eligible Leads", "Converted by Name", "Conversion %", "Still Driving Evidence", "Still Driving %"])
    ws.append(["Facebook Leads File", 2, 2, 1.0, 1, 0.5])
    ws.append(["Microsoft Teams - Onboarding Drivers", 1, 0, 0, 0, 0])

    ws = wb.create_sheet("Rules and Issues")
    ws.append([])
    ws.append(["Rules and Issues"])
    ws.append(["No guessing, no fuzzy matching, and no order-log claims."])
    ws.append([])
    ws.append(["Item", "Detail"])
    ws.append(["Conversion rule applied", "Exact normalized lead name equals exact normalized Xcelerator driver full name."])
    ws.append(["Phone evidence", "Phone match is secondary evidence and does not override the name rule."])
    wb.save(path)


def test_hr_recruiting_workbook_source_projects_only_aggregate_evidence(tmp_path) -> None:
    workbook_path = tmp_path / "hr-kpi.xlsx"
    _write_hr_kpi_workbook(workbook_path)

    dataset = asyncio.run(
        get_hr_recruiting_dataset(
            now=datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc),
            config=HrRecruitingConfig(workbook_path=str(workbook_path)),
        )
    )

    assert dataset["projection_mode"] == "read_only"
    assert dataset["source_profile"] == "kpi_workbook"
    assert dataset["source"] == "hr_kpi_workbook"
    assert dataset["source_system"] == "HR Lead KPI Recheck workbook"
    assert dataset["source_artifact"] == "hr-kpi.xlsx"
    assert dataset["summary"]["first_touch_eligible_count"] == 3
    assert dataset["summary"]["first_touch_within_24h_count"] == 1
    assert dataset["summary"]["first_touch_24h_pct"] == 0.3333
    assert dataset["summary"]["stale_untouched_48h"] == 1
    assert dataset["hard_targets"]["first_touch_24h_pct"]["status"] == "warning"
    assert dataset["hard_targets"]["new_hires_7d"]["status"] == "awaiting_feed"
    assert dataset["hard_targets"]["active_qualified_pipeline"]["status"] == "awaiting_feed"
    assert dataset["workbook_evidence"]["kpi_summary"]["total_outbound_attempts"] == 2
    assert dataset["workbook_evidence"]["source_log_qa"][0]["used_for_mapping"] is True
    team_members = {row["name"]: row for row in dataset["team_members"]}
    assert team_members["Jordan"]["status"] == "configured"
    assert team_members["Recruiter A"]["status"] == "source_backed"
    assert team_members["Recruiter A"]["first_outreach_leads"] == 2
    assert team_members["Recruiter A"]["real_discussion_leads"] == 2
    assert team_members["Recruiter A"]["within_24h_rate"] == 0.5
    assert dataset["row_counts"]["exception_queue_rows"] == 3
    exception_queue = dataset["workbook_evidence"]["exception_queue"]
    assert {row["category"] for row in exception_queue} == {
        "Failed No Outbound",
        "Failed Over 72h",
        "No Real Discussion",
    }
    assert {row["source_sheet"] for row in exception_queue} == {
        "Failed No Outbound",
        "Failed Over 72h",
        "No Real Discussion",
    }
    assert all(row["projection_mode"] == "read_only" for row in exception_queue)
    assert all(row["pii_suppressed"] is True for row in exception_queue)
    assert all(row["masked_contact"] == "PII suppressed" for row in exception_queue)
    assert all(row["lead_ref"].startswith("lead-ref-") for row in exception_queue)
    assert not re.search(r"(?:\+?1[\s.-]?)?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}", str(exception_queue))

    serialized = json.dumps(dataset)
    assert "Private Applicant" not in serialized
    assert "555-0100" not in serialized
    assert "one@example.com" not in serialized


def test_hr_recruiting_conversion_workbook_projects_masked_funnel(tmp_path) -> None:
    workbook_path = tmp_path / "hr-kpi.xlsx"
    conversion_path = tmp_path / "hr-conversion.xlsx"
    _write_hr_kpi_workbook(workbook_path)
    _write_hr_conversion_workbook(conversion_path)

    dataset = asyncio.run(
        get_hr_recruiting_dataset(
            now=datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc),
            config=HrRecruitingConfig(
                workbook_path=str(workbook_path),
                conversion_workbook_path=str(conversion_path),
            ),
            start_date="2026-05-14",
            end_date="2026-05-14",
        )
    )

    funnel = dataset["workbook_evidence"]["conversion_funnel"]
    assert funnel["projection_mode"] == "read_only"
    assert funnel["pii_suppressed"] is True
    assert funnel["source_status"] == "ok"
    assert funnel["workbook_name"] == "hr-conversion.xlsx"
    assert funnel["summary"] == {
        "eligible_leads": 2,
        "converted_leads": 1,
        "not_converted_leads": 1,
        "conversion_rate": 0.5,
        "still_driving_count": 1,
        "still_driving_rate": 0.5,
        "phone_match_count": 1,
        "unfiltered_eligible_leads": 3,
    }
    assert dataset["row_counts"]["conversion_source_rows"] == 3
    assert dataset["row_counts"]["conversion_eligible_leads"] == 2
    assert dataset["row_counts"]["conversion_converted_leads"] == 1
    assert funnel["trend_summary"]["current"]["eligible_leads"] == 2
    assert funnel["trend_summary"]["previous"]["eligible_leads"] == 1
    assert funnel["trend_summary"]["converted_change_pct"] == 0.0
    assert {row["source_bucket"] for row in funnel["by_source"]} == {
        "Facebook Leads File",
        "Microsoft Teams - Onboarding Drivers",
    }
    assert {row["date"] for row in funnel["trend"]} == {"2026-05-14"}

    serialized = json.dumps(dataset)
    assert "Private Converted" not in serialized
    assert "Private Not Converted" not in serialized
    assert "555-0100" not in serialized
    assert "DX-1" not in serialized


def test_hr_recruiting_workbook_filters_selected_range_and_compares_previous(tmp_path) -> None:
    workbook_path = tmp_path / "hr-kpi.xlsx"
    _write_hr_kpi_workbook(workbook_path)

    dataset = asyncio.run(
        get_hr_recruiting_dataset(
            now=datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc),
            config=HrRecruitingConfig(workbook_path=str(workbook_path)),
            start_date="2026-05-14",
            end_date="2026-05-14",
        )
    )

    assert dataset["date_range"] == {
        "start": "2026-05-14",
        "end": "2026-05-14",
        "days": 1,
        "previous_start": "2026-05-13",
        "previous_end": "2026-05-13",
    }
    assert dataset["source_status"] == "ok"
    assert dataset["row_counts"]["source_rows"] == 3
    assert dataset["row_counts"]["deduped_leads"] == 2
    assert dataset["row_counts"]["unfiltered_deduped_leads"] == 3
    assert dataset["row_counts"]["call_attempt_rows"] == 1
    assert dataset["row_counts"]["unfiltered_call_attempt_rows"] == 2
    assert dataset["period_metrics"] == {
        "new_leads": 2,
        "new_applicants": 2,
        "interviews_scheduled": 0,
        "new_hires": 0,
        "follow_ups": 1,
    }
    assert dataset["trend_comparison"]["previous"]["new_leads"] == 1
    assert dataset["trend_comparison"]["lead_volume_change_pct"] == 100.0
    assert dataset["workbook_evidence"]["kpi_summary"]["unique_lead_forms"] == 2
    assert dataset["workbook_evidence"]["kpi_summary"]["total_outbound_attempts"] == 1
    assert dataset["row_counts"]["exception_queue_rows"] == 2
    assert {row["category"] for row in dataset["workbook_evidence"]["exception_queue"]} == {
        "Failed No Outbound",
        "No Real Discussion",
    }

    serialized = json.dumps(dataset)
    assert "Private Applicant" not in serialized
    assert "555-0100" not in serialized
    assert "one@example.com" not in serialized


def test_hr_recruiting_workbook_range_uses_intake_date_not_modified_backlog(tmp_path) -> None:
    workbook_path = tmp_path / "hr-kpi.xlsx"
    _write_hr_kpi_workbook(workbook_path, include_modified_backlog=True)

    dataset = asyncio.run(
        get_hr_recruiting_dataset(
            now=datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc),
            config=HrRecruitingConfig(workbook_path=str(workbook_path)),
            start_date="2026-05-14",
            end_date="2026-05-14",
        )
    )

    assert dataset["row_counts"]["source_rows"] == 4
    assert dataset["row_counts"]["deduped_leads"] == 2
    assert dataset["period_metrics"]["new_leads"] == 2
    assert dataset["workbook_evidence"]["kpi_summary"]["unique_lead_forms"] == 2
    assert all(
        row["lead_created_date"] != "2026-03-20"
        for row in dataset["workbook_evidence"]["exception_queue"]
    )

    serialized = json.dumps(dataset)
    assert "Private Backlog Applicant" not in serialized
    assert "555-0199" not in serialized
    assert "backlog@example.com" not in serialized


def test_hr_recruiting_workbook_empty_current_range_preserves_prior_period(tmp_path) -> None:
    workbook_path = tmp_path / "hr-kpi.xlsx"
    _write_hr_kpi_workbook(workbook_path)

    dataset = asyncio.run(
        get_hr_recruiting_dataset(
            now=datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc),
            config=HrRecruitingConfig(workbook_path=str(workbook_path)),
            start_date="2026-05-15",
            end_date="2026-05-15",
        )
    )

    assert dataset["source_status"] == "empty"
    assert dataset["source_message"] == "HR KPI workbook is present but no Lead Level KPI rows matched the selected date range."
    assert dataset["row_counts"]["source_rows"] == 3
    assert dataset["row_counts"]["deduped_leads"] == 0
    assert dataset["row_counts"]["unfiltered_deduped_leads"] == 3
    assert dataset["period_metrics"]["new_leads"] == 0
    assert dataset["trend_comparison"]["current"]["new_leads"] == 0
    assert dataset["trend_comparison"]["previous"]["new_leads"] == 2
    assert dataset["trend_comparison"]["lead_volume_change_pct"] == -100.0
    assert dataset["by_worklist"] == []
    assert dataset["workbook_evidence"]["kpi_summary"]["unique_lead_forms"] == 0
    assert dataset["row_counts"]["exception_queue_rows"] == 0
    assert dataset["workbook_evidence"]["exception_queue"] == []


def test_hr_recruiting_workbook_path_wins_over_legacy_snapshot(tmp_path) -> None:
    workbook_path = tmp_path / "hr-kpi.xlsx"
    _write_hr_kpi_workbook(workbook_path)
    state_path = tmp_path / "hr-recruiting.json"
    state_path.write_text(
        json.dumps(
            {
                "rows": [
                    {
                        "source_email_id": "legacy-message",
                        "applicant": "Private Legacy Applicant",
                        "worklist": "Legacy Snapshot",
                        "status": "Assigned",
                        "first_assigned_at": "2026-05-14T10:00:00Z",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    dataset = asyncio.run(
        get_hr_recruiting_dataset(
            now=datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc),
            config=HrRecruitingConfig(
                workbook_path=str(workbook_path),
                snapshot_path=str(state_path),
                source="zapier_table",
            ),
        )
    )

    assert dataset["source_profile"] == "kpi_workbook"
    assert dataset["source_artifact"] == "hr-kpi.xlsx"
    assert dataset["row_counts"]["source_rows"] == 3
    assert {row["worklist"] for row in dataset["by_worklist"]} != {"Legacy Snapshot"}


def test_hr_recruiting_default_source_requires_workbook_path() -> None:
    dataset = asyncio.run(
        get_hr_recruiting_dataset(
            now=datetime(2026, 5, 15, 12, 0, tzinfo=timezone.utc),
            config=HrRecruitingConfig(),
        )
    )

    assert dataset["source"] == "hr_kpi_workbook"
    assert dataset["source_profile"] == "kpi_workbook"
    assert dataset["source_status"] == "snapshot_not_configured"
    assert "HR_RECRUITING_WORKBOOK_PATH" in str(dataset["source_message"])
